print("Loading Application...")


import time
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

from PlayActions import send_keys as send
from openpyxl import load_workbook
from AutocadBOM import AutocadBOM
from EVOBOM import EVOBOM
from EVOUtil import quitProgram
import xlrd

isAutoCad = False

def getInputs():
    root = tk.Tk()
    #root.iconbitmap(default="importBom.ico")
    root.withdraw()
    global isAutoCad
    
    topBOM = None
    
    # get the desired bom file
    
    file = filedialog.askopenfilename(title = "Select BOM File", initialdir="\\\\FS2\\engineer\\WORK\\Outdwg\\SolidworksBomOutputs")
    if file == "":
        quitProgram()
        
    # Checking for solidworks export file
    
    if file.endswith(".xlsm"):
        wb = load_workbook(file)
        ws = wb.active
        
        # read the last row of column A to get the top BOM number
        
        topBOM = ws["A" + str(ws.max_row)].value
        isAutoCad = False
        
        
    # Checking for autocad export file
    
    elif file.endswith(".xls"):
        wb = xlrd.open_workbook(file)
        ws = wb.sheet_by_index(0)
        isAutoCad = True
        
        slashIndex = file.rfind("/")
        if slashIndex == -1:
            slashIndex = 0
            
        partNumber = file[slashIndex: file.rfind(".")]
        partNumber = partNumber[1:5] + "-" + "0z" + "-" + partNumber[5:]
        
        topBOM = partNumber
        
        #return None, file
        
    else:
        messagebox.showerror("Error", "Invalid file type. Please select an Excel file.")
        quitProgram()
        
        

    if topBOM is None:
        print("Unable to auto detect top BOM number.")
        topBOM = simpledialog.askstring(title = "Top BOM Number", prompt = "What is the part number of the top BOM?")

    if topBOM is None:
        quitProgram()
    
    # proper topBOM format: 1234-5A-6789
    
    # certain checks to ensure the top BOM number is in the correct format
    
    while len(topBOM) != 12 or topBOM[4] != "-" or topBOM[7] != "-" or not topBOM[:4].isdigit() or not topBOM[5].isdigit() or not topBOM[6].isalpha() or not topBOM[8:].isdigit():
        messagebox.showerror("Error", "Invalid top BOM number. Please enter the top BOM number in the format 1234-5A-6789.\nYou entered: " + topBOM)
        
        topBOM = simpledialog.askstring(title = "Top BOM Number", prompt = "What is the part number of the top BOM?")
        
        if topBOM is None:
            quitProgram()
            
    
    return topBOM, file


topBOM, file = getInputs()


def runSolidworksUpload(topBOM, file):
    startTime = time.time()

    newBOM = EVOBOM(topBOM, file)

    newBOM.getAllParts()

    newBOM.writeBOMCSV()

    newBOM.createNew()

    newBOM.importBOM()

    newBOM.generateErrorReport()

    newBOM.getBadParts()


    timer = time.time()
    for item in newBOM.badParts:
        part = newBOM.findPart(item)
        if part is None:
            print("Part not found in BOM:", item)
            continue
        
        print("Creating part", part.partNumber, end="")
        part.createNew()
        timer = time.time()

    newBOM.loadBomCheck()

    # messagebox.showinfo("Verify", "Please verify the parts in the inventory before continuing.\nPress OK to continue.")
    BOMGood = messagebox.askquestion("Verify", "Please verify the parts in the inventory before continuing.\nUpload BOM to server?", icon = "warning")
    
    if BOMGood == "no":
        quitProgram()
    
    time.sleep(1)
    
    send([0.5, "alt x", 0.5, "alt x", 0.5])
 
    newBOM.uploadBOMToDatabase()

    totalTime = time.time() - startTime

    print("Complete", "BOM import has completed successfully.\nTotal time: " + str(totalTime/60) + " minutes.")
    for part in newBOM.badParts:
        print("Part not found in inventory:", part)
    if len(newBOM.badParts) == 0:
        print("No bad parts were found in the inventory.")

    finishMessage = f"""BOM import has completed successfully.

        Total time: {str(int(totalTime/60))} minutes and {str(int(totalTime%60))} seconds."""


    if len(newBOM.badParts) > 0:
        finishMessage += f"\n\n{len(newBOM.badParts)} parts were not found in the inventory and have been created:\n"
        for part in newBOM.badParts:
            finishMessage += f"{part}\n"
        

    messagebox.showinfo("Complete", finishMessage)

def runAutocadUpload(topBOM, file):
    newBOM = AutocadBOM(topBOM, file)
    newBOM.formatCSV()
    
if isAutoCad:
    print("Running AutoCad Upload...")
    runAutocadUpload(topBOM, file)
    
else:
    print("Running Solidworks Upload...")
    runSolidworksUpload(topBOM, file)
    
print("Application has finished running.")

messagebox.showinfo("Program Finished", "Program has finished.")