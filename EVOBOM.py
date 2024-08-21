import glob
import os
from PlayActions import send_keys as send
from openpyxl import load_workbook
import xlrd
from EVOPart import Part
from EVOUtil import createNewPart, isNone, openTASProgram, abreviateWords


class EVOBOM():
    def __init__(self, partNumber, filename):
        self.partNumber = partNumber.upper()
        self.filename = filename
        self.description = "Bill of Materials"
        
        self.partClass = "ELEC" if self.partNumber[5] == "9" else "MECH"
        self.partType = "A"
        
        
    def getAllParts(self):
        print("Reading BOM file", self.filename, end="...\n")
        
        wb = load_workbook(self.filename)
        ws = wb.active
        
        self.partsList = []
        self.description = ws["B" + str(ws.max_row)].value
        
        for row in range(ws.max_row - 2, 0, -1):
            itemNumber = str(ws["A" + str(row)].value)
            partNumber = str(ws["B" + str(row)].value)
            
            if partNumber == "None":
                continue
            
            description = str(ws["C" + str(row)].value).strip()
            specs = str(ws["D" + str(row)].value) + "\n" + str(ws["E" + str(row)].value).strip()
            vendor = str(ws["F" + str(row)].value).strip()
            vendorPartNumber = str(ws["G" + str(row)].value).strip()
            mfg = str(ws["H" + str(row)].value).strip()
            mfgPartNumber = str(ws["I" + str(row)].value).strip()
            referenced = (ws["J" + str(row)].value is not None and str(ws["J" + str(row)].value).strip() != "")
            print("Referenced: ", referenced, ws["J" + str(row)].value)
            
            partClass = str(ws["K" + str(row)].value).strip()
            partType = str(ws["L" + str(row)].value).strip()
            
            
            qty = str(ws["L" + str(row)].value)
            
            description = abreviateWords(description)
            specs = abreviateWords(specs)
            
            if isNone(specs):
                specs = ""
            if isNone(vendor):
                vendor = ""
            else:
                vendor = vendor[:4]+"50"
                
            if isNone(vendorPartNumber):
                vendorPartNumber = ""
            if mfg == "None":
                mfg = ""
            if isNone(mfgPartNumber):
                mfgPartNumber = ""
                
            if isNone(partClass):
                partClass = ""
            if isNone(partType):
                partType = ""
            
            print("Part Number: ", partNumber, "Referenced: ", referenced)
            if (partNumber is not None or partNumber.strip() != "") and not referenced:
                part = Part(partNumber.upper(), description, partClass= partClass, partType=partType, mfg=mfg, mfgNumber=mfgPartNumber, vendor=vendor, vendorNumber=vendorPartNumber, specs=specs)
                self.partsList.append((part, itemNumber, qty))
        
        print("...Finished reading BOM file\n")
                
    def getBadParts(self, reportFile = ""): # reportFile is in xls format
        print("Getting bad parts...")
        
        if reportFile == "":
            list_of_files = glob.glob("\\\\FS2\\engineer\\WORK\\Outdwg\\SolidworksBomOutputs\\*.xls")
            reportFile = max(list_of_files, key=os.path.getctime)
        
        
        wb = xlrd.open_workbook(reportFile)
        ws = wb.sheet_by_index(0)
        
        self.badParts = []
        for row in range(2, ws.nrows):
            if ws.cell_value(row, 0) == self.partNumber:
                self.badParts.append(ws.cell_value(row, 4))
        
        for part in self.badParts:
            print("New bad part found in BOM:", part)
            
        print("...Finished getting bad parts\n")
        
    def findPart(self, partNumber):
        for part in self.partsList:
            if part[0].partNumber == partNumber:
                return part[0]
            
        return None
        

    def writeBOMCSV(self):
        
        print("Writing BOM CSV...")
        BOMImportDir = "\\\\Erp\\dbamfg\\IMPORT\\BOM.csv"
        with open(BOMImportDir, "w") as file:
            file.write("Product Code (Part Number),Product Class,Type (NRMFABLTKO),Stock Unit of Measure\n")
            file.write("DO NOT REARRANGE THE ORDER OF THE COLUMNS IN THE FILE\n")
            
            for part in self.partsList:
                file.write(f"{self.partNumber},{part[1]},{part[0].partNumber},{part[2]}\n")
                
            file.close()
        
        print("...Finished writing BOM CSV\n")
    
    def importBOM(self):
        print("Importing BOM...")
        
        openTASProgram("ImportBom")
        send(["tab 2", "left", "enter 2", 5, "enter"]) # 3
        
        print("...Finished importing BOM\n")
        
    def writePartsCSV(self, badParts):
        PartsImportDir = "\\\\Erp\\dbamfg\\IMPORT\\INVENTORY.csv"
        with open(PartsImportDir, "w") as file:
            file.write("Product Code (Part Number),Product Class,Type (NRMFABLTKO),Stock Unit of Measure,Purchase Unit Measure,Sales Unit Measure,Description,Description Line 2,Drawing Number,Primary Vendor Code,Primary Vendor Item Number,Specs Line 1,Specs Line 2,Specs Line 3\n")
            file.write("DO NOT REARRANGE THE ORDER OF THE COLUMNS IN THE FILE\n")
            
            for i in badParts:
                part = self.findPart(i)
                file.write(f"{part.partNumber},{part.partClass},{part.partType},EA,EA,EA,{part.description[:30]},{part.description[30:]},{part.partNumber},{part.vendor},{part.vendorNumber},{part.specs[:30]},{part.specs[30:60]},{part.specs[60:]}\n")
            file.close()
    
    def importParts(self):
        openTASProgram("DBB")
        send(["tab 2", "left", "enter 2", 10, "enter"])    # 8
        
    def createNew(self):
        print("Creating new BOM...")
        createNewPart(self.partNumber, self.description, self.partClass, self.partType)
        print("...Finished creating new BOM\n")
        
            
    def generateErrorReport(self):
        print("Generating error report...")
        openTASProgram("BomErrorReport")
        
        send(["enter", self.partNumber, "enter 2", "alt p", 2, "alt o", 2])
        
        send(["alt f", "tab 2", "ctrl delete", "E", "tab", "E:\\WORK\\OUTDWG\\SOLIDWORKSBOMOUTPUTS\\", "alt o", 1,  "enter"])
        
        print("...Finished generating error report\n")

    def loadBomCheck(self):
        openTASProgram("EditBom")
        send([self.partNumber, "enter", 0.5])
        
    def uploadBOMToDatabase(self):
        send(["alt x", 1])
        openTASProgram("DCE")
        send(["n", "alt t", "enter", 2, "enter", 2])
        
        
"""
Input Format:
7/23/2024			Error Report for Imported Data
BILLS OF MATERIAL ERRORS												
Parent Part                 Component                         Quantity Per												
1001-1A-0049                7825-7D-3262		X-BKICM	      1.00000000					
1001-1A-0049				7825-9D-3261		X-BKICM	      1.00000000					
1001-1A-0057				7825-7D-3262		X-BKICM	      1.00000000					
1001-1A-0057				7825-9D-3261		X-BKICM	      1.00000000	


Output Format for parts:
Product Code (Part Number)	Product Class	Type (NRMFABLTKO)	Stock Unit of Measure	Purchase Unit Measure	Sales Unit Measure	Description	Description Line 2	Drawing Number	Primary Vendor Code	Primary Vendor Item Number	Specs Line 1	Specs Line 2	Specs Line 3
DO NOT REARRANGE THE ORDER OF THE COLUMNS IN THE FILE			
1001-1A-0065	            ELEC	        A	                EA	                    EA	                                                                        1001-1A-0065

Output format:
Parent Part Code    Line Number	    Component Part Code	    Quantity Required
DO NOT REARRANGE THE ORDER OF THE COLUMNS IN THE FILE			
1001-1A-0065        23	            7825-9A-2506	        3
1001-1A-0065	    22	            7855-9A-0649	        1
1001-1A-0065	    21	            7855-9A-0601	        1
				


"""

